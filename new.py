import openpyxl
import pandas as pd
# Load the workbook
workbook = openpyxl.load_workbook('C:\cricket_scores.xlsx')
# Get the active worksheet (scoreboard)
worksheet = workbook.active
# Calculate and update NRR for each team
for row in range(2, 6):
    matches_played = float(worksheet[f'B{row}'].value)
    total_runs_scored = float(worksheet[f'C{row}'].value)
    total_overs_faced = float(worksheet[f'D{row}'].value)
    total_runs_conceded = float(worksheet[f'E{row}'].value)
    total_overs_bowled = float(worksheet[f'F{row}'].value)
    net_run_rate = (total_runs_scored / total_overs_faced - total_runs_conceded / total_overs_bowled)
    worksheet[f'G{row}'] = net_run_rate


# Create a new sheet for NRR
nrr_sheet = workbook.create_sheet(title='NRR')

# Copy the NRR data from the scoreboard sheet
for row in range(1, 6):
    for col in range(1, 8):
        nrr_sheet.cell(row=row, column=col).value = worksheet.cell(row=row, column=col).value


# Create a new sheet for the points table
points_sheet = workbook.create_sheet(title='Points Table')

# Define the headers for the points table
headers = ['Team', 'Matches Played', 'Wins', 'Losses', 'Points']
for col, header in enumerate(headers, start=1):
    points_sheet.cell(row=1, column=col).value = header

# Calculate and populate the points table
for row in range(2, 6):
    team = worksheet[f'A{row}'].value
    matches_played = float(worksheet[f'B{row}'].value)
    wins = float(worksheet[f'C{row}'].value)
    losses = float(worksheet[f'D{row}'].value)
    points = wins * 2
    points_sheet.cell(row=row, column=1).value = team
    points_sheet.cell(row=row, column=2).value = matches_played
    points_sheet.cell(row=row, column=3).value = wins
    points_sheet.cell(row=row, column=4).value = losses
    points_sheet.cell(row=row, column=5).value = points

# Save the changes to the workbook
workbook.save('C:\cricket_scores.xlsx')